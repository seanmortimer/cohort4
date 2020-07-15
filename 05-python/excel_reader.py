from openpyxl import load_workbook
print()

def load_excelbook(file):
    wb = load_workbook(file)
    wb_dict = {}  

    # for sheet_name in wb.sheetnames:
        # wb_dict[sheet_name] = {} 
    # print(wb_dict)

    for sheet in wb.worksheets:
        # print('sheet:', sheet.title)
    #     # print(sheet)
        wb_dict[sheet.title] = {}


        for row in sheet.iter_rows(min_row=2, min_col=1):
            id = row[0].value
            wb_dict[sheet.title][id] = {}
            for cell in row:
                label = sheet.cell(row=1, column=cell.column).value
                value = cell.value
                # print(label)
                # print(value)
                wb_dict[sheet.title][id][label] = value
                # print('wb_dict[sheet.title]:', wb_dict[sheet.title])
    #         #     col = 0
    #         #     for cell in row:
    #         #         wb_dict[sheet][i][wb[sheet][2][col].value] = cell.value
    #         #         col += 1
    #         # i += 1
    return(wb_dict)
