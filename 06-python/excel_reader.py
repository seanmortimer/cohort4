from openpyxl import load_workbook
print()

# ADD ONS: Data validation,
# make into class?
# make print width more adjustable


def excelToDict(file):
    wb = load_workbook(file)
    wb_dict = {}

    for sheet in wb.worksheets:
        wb_dict[sheet.title] = {}

        for row in sheet.iter_rows(min_row=2, min_col=1):
            id = row[0].value
            wb_dict[sheet.title][id] = {}
            for cell in row:
                label = sheet.cell(row=1, column=cell.column).value
                value = cell.value
                wb_dict[sheet.title][id][label] = value

    return wb_dict


def alternateWay(file):
    wb = load_workbook(file)
    wb_dict = {}

    for sheet in wb.worksheets:
        wb_dict[sheet.title] = {}
        i = 1
        for row in sheet.values:
            if i == 1:
                labels = row
            else:
                wb_dict[sheet.title][row[0]] = dict(zip(labels, row))
            i += 1

    return wb_dict


def buildInvoice(file, invoice_no):
    wb_dict = excelToDict(file)
    invoice = {"invoice_no": invoice_no}

    custID = wb_dict["Invoices"][invoice_no]["customer_id"]
    invoice["customer"] = wb_dict["Customers"][custID]

    # invoice["invoice_no"] = invoice_no
    invoice["date"] = wb_dict["Invoices"][invoice_no]["date"]
    invoice["payment"] = wb_dict["Invoices"][invoice_no]["payment"]
    items = [value for (line, value) in wb_dict["Line_Items"].items() if value["invoice_no"] == invoice_no]
    invoice["line_items"] = {}
    invoice["total"] = 0

    for item in items:
        product = wb_dict["Products"][item["product_id"]]
        unit_price = product["price"]
        qty = item["quantity"]
        price = unit_price * qty
        line = len(invoice["line_items"]) + 1

        invoice["line_items"][line] = {
            "product_id": item["product_id"],
            "name": product["product_name"],
            "unit_price": unit_price,
            "qty": int(qty),
            "price": price
        }
        invoice["total"] += price

    return invoice


def printInvoice(file, invoice_no):
    print_width = 80
    invoice = buildInvoice(file, invoice_no)

    i_str = f"Invoice No: {invoice_no}"
    date = f"Date: {invoice['date']}"
    name = f"{invoice['customer']['first_name']} {invoice['customer']['last_name']}"
    addr = f"{invoice['customer']['address']}"
    phone = invoice["customer"]["phone"]

    # header
    print("*".ljust(print_width, '*'))
    print("\n" + name.ljust(print_width - len(i_str)) + i_str)
    print(addr + date.rjust(print_width - len(addr)))
    print(f"{phone}\n\n")

    # body
    print("-".ljust(print_width, '-'))
    print("Item No".ljust(18), "Name".ljust(25),
          "Unit Price".ljust(14), "Quantity".ljust(13), "Price")
    print("-".ljust(print_width, '-'))

    for line in invoice['line_items'].values():
        print(
            f"  {int(line['product_id'])}".ljust(10),
            f"{line['name']}"[0:60].ljust(35),
            f"${line['unit_price']:.2f}",
            f"{line['qty']}".rjust(12),
            f"$ {line['price']:.2f}".rjust(12))

    print("-".ljust(print_width, '-'))
    print(f"$ {invoice['total']:.2f}".rjust(print_width - 1))

    # footer
    print()
    print("Thank you for choosing Billy's Grocery!".center(print_width))
    print("*".ljust(print_width, '*'))

    return invoice['total']


# printInvoice('./data/billy_test_data.xlsx', 1)
# printInvoice('./data/billy_accounting.xlsx', 2)
